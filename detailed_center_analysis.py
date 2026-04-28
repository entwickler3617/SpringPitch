#!/usr/bin/env python3
"""
스프링 중심축 정밀 분석 및 rθz 좌표계 기준점 도출

프로세스(이미지/사양서 기준) 반영:
    1) IGS 포인트 로드
    2) IGS 포인트 보정(축 정렬/곡률 평활화 - 경량 구현)
    3) 좌표변환 xyz -> rθz (축: 상/하단 센터를 잇는 축 ≈ PCA 주축, 원점: 상단 센터 기본)
    4) 좌표변환 테이블 기반 기초 치수 산출 (Turn, Radius, Height, Pitch 등)
    5) 3점 호 보간 기반 정밀 치수 산출은 제외 (요청사항에 따라 계산결과 산출 스킵)
"""

import re
import numpy as np
import matplotlib
matplotlib.use('Agg')  # Use non-GUI backend for headless execution
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
import os
import math
import csv
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl import Workbook, load_workbook
from typing import Tuple, Any
import logging
import pandas as pd
import shutil
import sys
import glob

# 정상 동작 기본값: 선형(폴리라인) 등간격 표준화로 형상 보존
# 환경변수 SPRING_NORM_METHOD 로 변경 가능 (linear|uniform_bspline|nurbs)
NORMALIZATION_METHOD = os.getenv('SPRING_NORM_METHOD', 'linear').lower()
BSPLINE_DEGREE_DEFAULT = int(os.getenv('SPRING_BSPLINE_DEGREE', '3'))
BSPLINE_SMOOTHING_DEFAULT = float(os.getenv('SPRING_BSPLINE_SMOOTHING', '0.0'))  # splprep의 s 값 (0이면 보간에 가깝고, >0이면 스무딩)
END_SLICE_PERCENT_DEFAULT = float(os.getenv('SPRING_END_SLICE_PERCENT', '5.0'))
NURBS_MAXFIT_DEFAULT = int(os.getenv('SPRING_NURBS_MAXPTS', '300'))
# 시작/종료 선제 보정 모드: 'none' | 'min_gap' | 'complete_turns' | 'min_gap+complete_turns'
SEAM_MODE = os.getenv('SPRING_SEAM_MODE', 'none').lower()
# θ 내보내기 모드: 'raw_wrapped' | 'raw_unwrapped' | 'start0_unwrapped'
# 기본값을 'start0_unwrapped'로 하여 시작점 θ=0도를 보장
THETA_MODE = os.getenv('SPRING_THETA_MODE', 'start0_unwrapped').lower()
# 단조 증가 방향 강제 (시계/반시계 중 양(+) 증가로 맞춤). '1'/'true' 활성화
THETA_POSITIVE = os.getenv('SPRING_THETA_POSITIVE', 'true').strip().lower() in ('1','true','yes','on')
ARC_PLANE_MODE = os.getenv('SPRING_ARC_PLANE', 'local_xy').lower()
START_STRATEGY = os.getenv('SPRING_START_STRATEGY', 'min_radius').lower()  # min_radius|min_z|max_z|auto
SMOOTH_MODE = os.getenv('SPRING_SMOOTH_MODE', 'none').lower()            # none|robust_ends|robust_all
PRESERVE_RAW_XYZ = os.getenv('SPRING_PRESERVE_RAW_XYZ', 'true').strip().lower() in ('1','true','yes','on')
STD_MODE = os.getenv('SPRING_STD_MODE', 'local').strip().lower()  # 레거시: local | raw_preserve | raw_with_outlier_fix
EFG_MODE = os.getenv('SPRING_EFG_MODE', os.getenv('SPRING_STD_MODE', 'local')).strip().lower()
OUTLIER_ENABLE = os.getenv('SPRING_OUTLIER_ENABLE', 'true').strip().lower() in ('1','true','yes','on')
OUTLIER_RADIUS_MAD = float(os.getenv('SPRING_OUTLIER_RADIUS_MAD', '3.5'))
OUTLIER_STEP_MAD = float(os.getenv('SPRING_OUTLIER_STEP_MAD', '4.0'))
ALIGN_START = os.getenv('SPRING_ALIGN_START', 'true').strip().lower() in ('1','true','yes','on')
# 시작 원점 방식: start(요청사항: 시작점에서 Z≈0 되도록) | midpoint(기존 반대위상 중점)
ORIGIN_MODE = os.getenv('SPRING_ORIGIN_MODE', 'start').strip().lower()  # start | midpoint
# 단순 직교→원통 변환 전용 모드: x_norm,y_norm,z_norm만으로 R,θ,Z 산출
SIMPLE_CONVERT = os.getenv('SPRING_SIMPLE_CONVERT', 'true').strip().lower() in ('1','true','yes','on')
SIMPLE_REBASE_Z0 = os.getenv('SPRING_SIMPLE_REBASE_Z0', 'true').strip().lower() in ('1','true','yes','on')
CYL_ONLY = os.getenv('SPRING_CYL_ONLY', 'false').strip().lower() in ('1','true','yes','on')
# B:D(x,y,z) 출력 모드: raw | local | sinusoidal (X=theta_unwrap_deg, Y=x_local, Z=y_local)
XYZ_MODE = os.getenv('SPRING_XYZ_MODE', 'raw').strip().lower()
# 끝단(초/말) 구간에서 이상치 판단을 강화하기 위한 계수(작을수록 민감). 예: 0.7
OUTLIER_END_FACTOR = float(os.getenv('SPRING_OUTLIER_END_FACTOR', '0.7'))
OUTLIER_END_PERCENT = float(os.getenv('SPRING_OUTLIER_END_PERCENT', '3.0'))
# Vectorial diameter: pointer offset in turns (default 0.5 turn = opposite phase)
VECTORIAL_TURN_OFFSET = float(os.getenv('SPRING_VECTORIAL_TURN_OFFSET', '0.5'))
# min_Pitch: use only rows with V>0 (Excel exact MIN(W) if false)
MIN_PITCH_POSITIVE_ONLY = os.getenv('SPRING_MIN_PITCH_POSITIVE_ONLY', 'false').strip().lower() in \
    ('1', 'true', 'yes', 'on')

def parse_igs_points(filepath):
    """IGS 파일에서 좌표점들을 추출 (섹션 기반 파서 사용, 단위 mm 정규화)."""
    try:
        from iges_parser import parse_iges_points as _parse_iges
        pts, glb = _parse_iges(filepath)
        return np.array(pts, dtype=float)
    except Exception as e:
        logging.warning(f"섹션 파서 실패({e}). 간단 정규식 방식으로 재시도합니다.")
        # 고급 수식 적용 (여러 타입)
        pattern = r"116,\s*([+-]?(?:\d+\.?\d*|\d*\.\d+)(?:[eEdD][+-]?\d+)?),\s*" \
                  r"([+-]?(?:\d+\.?\d*|\d*\.\d+)(?:[eEdD][+-]?\d+)?),\s*" \
                  r"([+-]?(?:\d+\.?\d*|\d*\.\d+)(?:[eEdD][+-]?\d+)?)(?:,\s*0)?\s*;"
        points = []  # 누락되었던 리스트 초기화
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as file:
            for line in file:
                line = line.strip()
                m = re.search(pattern, line)
                if m:
                    x = float(m.group(1).replace('D','E').replace('d','e'))
                    y = float(m.group(2).replace('D','E').replace('d','e'))
                    z = float(m.group(3).replace('D','E').replace('d','e'))
                    points.append([x, y, z])
        return np.array(points, dtype=float)

def find_spring_axis_pca(points):
    """PCA를 사용하여 스프링의 주축 찾기"""
    # 점들을 중심으로 이동
    center = np.mean(points, axis=0)
    centered_points = points - center
    
    # 공분산 행렬 계산
    cov_matrix = np.cov(centered_points.T)
    
    # 고유값과 고유벡터 계산
    eigenvalues, eigenvectors = np.linalg.eigh(cov_matrix)
    
    # 고유값에 따라 정렬 (큰 순서대로)
    idx = np.argsort(eigenvalues)[::-1]
    eigenvalues = eigenvalues[idx]
    eigenvectors = eigenvectors[:, idx]
    
    # 주축 (가장 큰 고유값에 해당하는 고유벡터)
    principal_axis = eigenvectors[:, 0]
    
    return center, principal_axis, eigenvalues, eigenvectors

def build_local_frame(axis_direction: np.ndarray):
    """축 방향으로부터 국소 좌표계(Ex,Ey,Ez)를 구성 (Ez = axis_direction).
    잘못 삽입되었던 엑셀 수식 관련 코드 제거.
    반환: (R, Ex, Ey, Ez) where R rows are basis vectors.
    """
    Ez = axis_direction / np.linalg.norm(axis_direction)
    ref = np.array([1.0, 0.0, 0.0])
    if abs(np.dot(Ez, ref)) > 0.99:
        ref = np.array([0.0, 1.0, 0.0])
    Ex = np.cross(ref, Ez)
    nEx = np.linalg.norm(Ex)
    if nEx < 1e-12:
        ref = np.array([0.0, 1.0, 0.0])
        Ex = np.cross(ref, Ez)
        nEx = np.linalg.norm(Ex)
    Ex /= nEx
    Ey = np.cross(Ez, Ex)
    Ey /= np.linalg.norm(Ey)
    R = np.vstack([Ex, Ey, Ez])
    return R, Ex, Ey, Ez

def to_local_coordinates(points: np.ndarray, origin: np.ndarray,
                        axis_direction: np.ndarray, logger=None) -> Tuple[np.ndarray, Any]:
    """글로벌 points를 주축과 원점 기준 국소 직교좌표(x,y,z)로 변환."""
    if logger is None:
        logger = logging.getLogger()

    if points is None or len(points) == 0:
        return np.zeros((0,3), dtype=float), logger

    # The primary axis for the spring is now X.
    # We need to define the local coordinate system accordingly.
    # Ez will be the spring's axis. Let's align it with the global X-axis for convention.
    # But the user wants the primary axis to be X_local.
    # So, the axis_direction corresponds to the local X-axis.
    
    Ex = axis_direction / np.linalg.norm(axis_direction)
    
    # Create a perpendicular vector for Y.
    # To ensure stability, choose a reference vector that is not parallel to Ex.
    ref = np.array([0.0, 0.0, 1.0])
    if abs(np.dot(Ex, ref)) > 0.99:
        ref = np.array([0.0, 1.0, 0.0])
        
    Ez = np.cross(Ex, ref)
    Ez /= np.linalg.norm(Ez)
    
    Ey = np.cross(Ez, Ex)
    Ey /= np.linalg.norm(Ey)

    rel = points - origin
    x = rel @ Ex
    y = rel @ Ey
    z = rel @ Ez
    
    logger.info("\n[to_local_coordinates] Conversion Summary:")
    logger.info(f"  - Local X (Primary): min={np.min(x):.3f}, max={np.max(x):.3f}, mean={np.mean(x):.3f}")
    logger.info(f"  - Local Y: min={np.min(y):.3f}, max={np.max(y):.3f}, mean={np.mean(y):.3f}")
    logger.info(f"  - Local Z: min={np.min(z):.3f}, max={np.max(z):.3f}, mean={np.mean(z):.3f}")
    
    return np.vstack([x,y,z]).T, logger

def cylindrical_from_local(local_xyz: np.ndarray) -> tuple[np.ndarray, np.ndarray, np.ndarray]:
    """로컬 직교좌표 -> (r, theta, z)."""
    if local_xyz is None or len(local_xyz) == 0:
        return np.zeros(0), np.zeros(0), np.zeros(0)
    x = local_xyz[:,0]; y = local_xyz[:,1]; z = local_xyz[:,2]
    r = np.hypot(x, y)
    theta = np.arctan2(y, x)
    return r, theta, z

def unwrap_theta(theta: np.ndarray, sort_idx: np.ndarray | None = None) -> np.ndarray:
    """theta를 연속적으로 언랩. sort_idx 제공 시 해당 순서로 정렬 후 언랩한 뒤 원래 순서로 복원."""
    if theta is None or len(theta) <= 1:
        return theta
    if sort_idx is None:
        return np.unwrap(theta)
    ordered = theta[sort_idx]
    unwrapped_ordered = np.unwrap(ordered)
    inv = np.empty_like(sort_idx)
    inv[sort_idx] = np.arange(len(sort_idx))
    return unwrapped_ordered[inv]

def moving_average(x: np.ndarray, window: int) -> np.ndarray:
    """단순 이동평균 (edge 패딩)."""
    if x is None or len(x) == 0:
        return np.zeros_like(x)
    window = max(1, int(window))
    if window <= 1 or window >= len(x):
        return x.astype(float)
    pad = window // 2
    xpad = np.pad(x.astype(float), (pad, pad), mode='edge')
    kernel = np.ones(window, dtype=float) / window
    y = np.convolve(xpad, kernel, mode='same')
    return y[pad:-pad]

def compute_basic_metrics(local_xyz: np.ndarray, r: np.ndarray, theta: np.ndarray, z: np.ndarray):
    """프로세스 4단계: 기초 치수 계산(보간 제외).
    - Turn: theta_unwrapped / (2*pi)
    - Height: z (상대값) 및 전체 높이
    - Pitch(local): dZ / dTurn (차분 기반), 이동평균으로 평활화
    - Diameter(Perpendicular): 2*r (기초값, 정밀 보간 제외)
    - Radius: r
    - Diameter(Vectorial): 기초 단계에서는 2*r와 동일하게 둠 (정밀 해석은 5단계)
    반환: dict of numpy arrays + summary
    """
    # z' 기준 정렬 인덱스
    sort_idx = np.argsort(z)
    theta_unwrapped = unwrap_theta(theta, sort_idx)
    turn = theta_unwrapped / (2.0 * np.pi)

    # 차분으로 local pitch 계산 (dZ/dTurn)
    dturn = np.diff(turn)
    dz = np.diff(z)
    with np.errstate(divide='ignore', invalid='ignore'):
        pitch_local = np.zeros_like(z)
        p = np.where(np.abs(dturn) > 1e-6, dz / dturn, np.nan)
        # 중앙에 배치 (i -> between i and i+1). 간단히 앞쪽에 복사/뒤쪽에 보간
        pitch_local[1:] = p
        pitch_local[0] = pitch_local[1]

    # 평활화(곡률/피치 부드럽게)
    r_smooth = moving_average(r, window= nine_if_small(len(r)))
    z_smooth = moving_average(z, window= nine_if_small(len(z)))
    pitch_smooth = moving_average(pitch_local, window= nine_if_small(len(pitch_local)))

    diameter_perp = 2.0 * r_smooth
    diameter_vec = 2.0 * r_smooth  # 정밀 차이는 5단계에서 처리 (현재 동일 처리)

    summary = {
        'height_total': float(np.nanmax(z) - np.nanmin(z)),
        'pitch_mean': float(np.nanmean(pitch_smooth)),
        'pitch_min_estimate': float(np.nanmin(pitch_smooth)),  # 참고값(보간 미적용)
        'turn_total_estimate': float((np.nanmax(turn) - np.nanmin(turn)))
    }

    cols = {
        'x_local': local_xyz[:, 0],
        'y_local': local_xyz[:, 1],
        'z_local': z,
        'r': r,
        'theta': theta,
        'theta_unwrapped': theta_unwrapped,
        'turn': turn,
        'r_smooth': r_smooth,
        'z_smooth': z_smooth,
        'pitch_local': pitch_local,
        'pitch_smooth': pitch_smooth,
        'diameter_perpendicular': diameter_perp,
        'diameter_vectorial': diameter_vec,
    }
    return cols, summary

def nine_if_small(n: int):
    """데이터 길이에 따라 이동평균 윈도우를 9 또는 n//50*2+1 등으로 설정."""
    w = max(5, (n // 50) * 2 + 1)
    if w % 2 == 0:
        w += 1
    return min(w, 101)  # 과도하게 크지 않도록 상한

def project_points_to_axis(points, axis_point, axis_direction):
    """점들을 축에 투영하여 축 상의 위치 계산"""
    # 축 방향 벡터 정규화
    axis_direction = axis_direction / np.linalg.norm(axis_direction)
    
    # 각 점에서 축 점까지의 벡터
    to_points = points - axis_point
    
    # 축 방향으로의 투영 길이
    projections = np.dot(to_points, axis_direction)
    
    # 축 상의 점들
    axis_points = axis_point + projections[:, np.newaxis] * axis_direction
    
    # 축으로부터의 거리
    distances = np.linalg.norm(points - axis_points, axis=1)
    
    return projections, distances, axis_points

def analyze_spring_layers(points, axis_point, axis_direction, num_layers=20):
    """스프링을 여러 층으로 나누어 각 층의 중심 분석"""
    projections, distances, axis_points = project_points_to_axis(points, axis_point, axis_direction)
    
    # 투영 길이 범위
    min_proj, max_proj = np.min(projections), np.max(projections)
    
    # 층별로 나누기
    layer_edges = np.linspace(min_proj, max_proj, num=num_layers + 1)
    layer_centers = []
    layer_radii = []
    layer_z_values = []
    
    for i in range(num_layers):
        # 해당 층에 속하는 점들
        mask = (projections >= layer_edges[i]) & (projections < layer_edges[i + 1])
        if np.sum(mask) > 0:
            layer_points = points[mask]
            layer_center = np.mean(layer_points, axis=0)
            layer_center_proj = (layer_edges[i] + layer_edges[i + 1]) / 2
            layer_center_on_axis = axis_point + layer_center_proj * axis_direction
            
            # 해당 층의 반지름들
            layer_distances = distances[mask]
            mean_radius = np.mean(layer_distances)
            
            layer_centers.append(layer_center)
            layer_radii.append(mean_radius)
            layer_z_values.append(layer_center[2])
    
    return np.array(layer_centers), np.array(layer_radii), np.array(layer_z_values)

def find_optimal_spring_center(points):
    """스프링의 최적 중심축 찾기"""
    
    # 방법 1: PCA를 사용한 주축 찾기
    pca_center, pca_axis, eigenvalues, eigenvectors = find_spring_axis_pca(points)
    
    # 스프링 층별 분석
    layer_centers, layer_radii, layer_z = analyze_spring_layers(points, pca_center, pca_axis)
    
    # Z축 방향 분석 (스프링 맨 위의 중앙 찾기)
    z_max = np.max(points[:, 2])
    z_min = np.min(points[:, 2])
    
    # 상위 10% 지점의 점들
    top_threshold = z_max - (z_max - z_min) * 0.1
    top_mask = points[:, 2] >= top_threshold
    top_points = points[top_mask]
    
    if len(top_points) > 0:
        top_center_xy = np.mean(top_points[:, :2], axis=0)
        top_center = np.array([top_center_xy[0], top_center_xy[1], z_max])
    else:
        top_center = np.array([pca_center[0], pca_center[1], z_max])
    
    # 상위 몇 개 층의 중심들 평균
    refined_top_center = top_center  # 기본값 설정
    if len(layer_centers) >= 3:
        top_3_layers = layer_centers[-3:]  # 상위 3개 층
        top_layers_center_xy = np.mean(top_3_layers[:, :2], axis=0)
        refined_top_center = np.array([top_layers_center_xy[0], top_layers_center_xy[1], z_max])
    
    return {
        'pca_center': pca_center,
        'pca_axis': pca_axis,
        'eigenvalues': eigenvalues,
        'top_center': top_center,
        'refined_top_center': refined_top_center,
        'layer_info': {
            'centers': layer_centers,
            'radii': layer_radii,
            'z_values': layer_z
        }
    }

def ensure_dir(path: str):
    if not os.path.isdir(path):
        os.makedirs(path, exist_ok=True)

def safe_save_workbook(wb, path: str, logger: logging.Logger | None = None):
    """Safely save an openpyxl workbook by writing to a temp file then atomically replacing.
    Prevents corrupted .xlsx if a crash occurs during save.
    """
    tmp_path = None
    try:
        d = os.path.dirname(path) or "."
        base = os.path.basename(path)
        tmp_path = os.path.join(d, f".~{base}.tmp")
        # Ensure any stale temp is removed
        if os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except Exception:
                pass
        wb.save(tmp_path)
        os.replace(tmp_path, path)
        if logger:
            logger.debug(f"Workbook safely saved to {path}")
    except Exception as e:
        if logger:
            logger.warning(f"안전 저장 실패: {e}")
        # Best-effort cleanup of temp
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except Exception:
                pass
        # Re-raise to allow caller to decide handling
        raise

def write_csv(path: str, header: list[str], columns: dict[str, np.ndarray]):
    n = len(next(iter(columns.values()))) if columns else 0
    with open(path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(header)
        for i in range(n):
            row = [columns[h][i] for h in header]
            writer.writerow(row)

import sys
# ... existing code ...
def write_excel_tk1(path: str, sheet_name: str, std_points: np.ndarray, raw_mapped: np.ndarray,
                    r_std: np.ndarray, theta_std_unwrapped: np.ndarray, z_std: np.ndarray,
                    extra_cols: dict | None = None,
                    summary_block: dict | None = None,
                    summary_grid: list[list] | None = None,
                    aux_red_headers: list[str] | None = None,
                    formula_from_aux: dict | None = None,
                    formula_spec: dict | None = None):
    """TK1.xlsx 산출 시트에 지정 컬럼만 안전하게 초기화/갱신한다.
    - 기존 파일/시트가 있어도 다른 컬럼/서식은 건드리지 않음
    - 대상 컬럼: [No, x, y, z, x_norm, y_norm, z_norm, R, θ, Z]
    - 헤더는 1행에, 데이터는 2행부터 추가.
    """
    headers = ['No', 'x', 'y', 'z', 'x_norm', 'y_norm', 'z_norm', 'R', 'θ', 'Z']
    n = len(std_points)

    # 1) 워크북 열기 (기존 파일이 있어야 함)
    if not os.path.exists(path):
        raise FileNotFoundError(f"TK1.xlsx not found at {path}")
    wb = load_workbook(path)

    # 2) 시트 사용: 기본은 기존 시트만 사용, 옵션에 따라 없으면 생성
    allow_create_sheet = os.getenv('SPRING_CREATE_ZERO1_IF_MISSING', 'false').strip().lower() in ('1','true','yes','on')
    if sheet_name not in wb.sheetnames:
        if allow_create_sheet:
            ws = wb.create_sheet(title=sheet_name)
            # 새로 만든 경우에만 최소 헤더를 기록 (서식 변경 없이 값만)
            for col_idx, head in enumerate(headers, start=1):
                ws.cell(row=1, column=col_idx, value=head)
            start_row = 2
        else:
            raise ValueError(f"Sheet '{sheet_name}' not found in {path}")
    else:
        ws = wb[sheet_name]
        # ws.max_row는 서식만 있어도 크게 나올 수 있어 신뢰하지 않음
        # 2행부터 핵심 컬럼들(A,B,E)에 값이 있는 마지막 행을 찾아 +1로 시작
        def find_first_empty_row(ws_, start=2, check_cols=(1, 2, 5)):
            r = start
            # 안전장치: 1백만 행을 넘지 않도록 제한
            limit = 1_000_000
            last_data_row = start - 1
            while r < limit:
                any_val = False
                for c in check_cols:
                    v = ws_.cell(row=r, column=c).value
                    if v is not None and str(v).strip() != "":
                        any_val = True
                        break
                if any_val:
                    last_data_row = r
                    r += 1
                else:
                    break
            return (last_data_row + 1) if last_data_row >= (start) else start
        start_row = find_first_empty_row(ws)

    # 3) 헤더는 기존에 있다고 가정 (단, 시트를 새로 만든 경우에만 위에서 기록함)

    # 4) 데이터 기록 (대상 컬럼만)
    for i in range(n):
        rrow = start_row + i
        ws.cell(rrow, 1, rrow - 1)  # No (starts from 1 when starting at row 2)
        # x,y,z는 원시(글로벌) 좌표 기록
        ws.cell(rrow, 2, float(raw_mapped[i, 0]))
        ws.cell(rrow, 3, float(raw_mapped[i, 1]))
        ws.cell(rrow, 4, float(raw_mapped[i, 2]))
        ws.cell(rrow, 5, float(std_points[i, 0]))
        ws.cell(rrow, 6, float(std_points[i, 1]))
        ws.cell(rrow, 7, float(std_points[i, 2]))
        ws.cell(rrow, 8, float(r_std[i]))
        ws.cell(rrow, 9, float(theta_std_unwrapped[i]))
        ws.cell(rrow, 10, float(z_std[i]))

    # 5) 추가 컬럼(L..): extra_cols가 제공되면, L열(12)부터 순서대로 기록
    header_to_col_letter: dict[str, str] = {}
    core_header_to_col_letter: dict[str, str] = {h: get_column_letter(i) for i, h in enumerate(headers, start=1)}
    if extra_cols:
        base_col = 12  # L
        col_idx = base_col
        for key, arr in extra_cols.items():
            # 헤더는 기존에 있다고 가정 (작성하지 않음)
            header_to_col_letter[key] = get_column_letter(col_idx)
            
            if arr is not None:
                for i in range(n):
                    rrow = start_row + i
                    ws.cell(rrow, col_idx,
                             None if len(arr) <= i or arr[i] is None or
                             (isinstance(arr[i], float) and (np.isnan(arr[i]) or np.isinf(arr[i])))
                             else float(arr[i]))
            # 폭은 기존 포맷 유지
            col_idx += 1
    
    # 6) 보기 편한 폭은 기존 포맷 유지

    # 7) 요약 블록(Z3:AC14) 기록
    if summary_grid:
        # summary_grid는 12행 x 4열 가정. Z3..AC14에 그대로 기록
        summary_start_row = 3
        summary_start_col = 26  # Z
        for i, row_vals in enumerate(summary_grid):
            for j, val in enumerate(row_vals):
                ws.cell(summary_start_row + i, summary_start_col + j, val)
    elif summary_block:
        base_row = 3
        r = base_row
        for label, value in summary_block.items():
            ws.cell(r, 26, label)  # Z
            ws.cell(r, 27, value)  # AA
            r += 1

    # L:X 헤더를 std 시트의 규격과 동일하게 하드코딩(환경 독립)
    # 순서: L(12)~X(24). None은 빈 헤더(셀 값 비움)를 의미
    std_lx_headers = [
        None,                         # L
        None,                         # M
        'Turn',                       # N
        'Radius',                     # O
        None,                         # P
        'Diameter (Perpendicular)',   # Q
        None,                         # R
        'Diameter (Vectorial)',       # S
        None,                         # T
        'Height',                     # U
        None,                         # V
        'Pitch',                      # W
        'min_Pitch',                  # X
    ]
    try:
        for idx, col in enumerate(range(12, 25)):
            ws.cell(row=1, column=col, value=std_lx_headers[idx])
    except Exception as e:
        logging.getLogger().warning(f"L:X 헤더 하드코딩 실패: {e}")

    # 최종 저장은 안전 저장으로 처리 (원본 파일 손상 방지)
    safe_save_workbook(wb, path, logging.getLogger())

def write_excel_raw(path: str, sheet_name: str, raw_points: np.ndarray):
    """Write all original raw IGS points (Step1) to a dedicated sheet.
    Columns: No, x, y, z. Clears existing data rows (keeps formatting).
    """
    if not os.path.exists(path):
        raise FileNotFoundError(f"TK1.xlsx not found at {path}")
    wb = load_workbook(path)
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet_name)
        headers = ['No', 'x', 'y', 'z']
        for i, h in enumerate(headers, start=1):
            ws.cell(row=1, column=i, value=h)
    else:
        ws = wb[sheet_name]
        # Clear existing data rows (2..max)
        for r in range(2, ws.max_row + 1):
            for c in range(1, 5):
                cell = ws.cell(row=r, column=c)
                if cell.value is not None and str(cell.value).strip() != "":
                    cell.value = None
    # Write raw points
    for i, p in enumerate(raw_points, start=2):
        ws.cell(i, 1, i - 1)
        ws.cell(i, 2, float(p[0]))
        ws.cell(i, 3, float(p[1]))
        ws.cell(i, 4, float(p[2]))
    safe_save_workbook(wb, path, logging.getLogger())

def visualize_spring_analysis(analysis_result, output_dir='output', std_df=None):
    """
    Generates a comprehensive visualization of the spring analysis results.
    """
    # Unpack data from the result dictionary
    local_coords = analysis_result['local_coords']
    resampled_points = analysis_result['resampled_points']
    r = analysis_result['r']
    z = analysis_result['z']
    theta_unwrapped_deg = analysis_result['theta_unwrapped_deg']
    outlier_mask = analysis_result['outlier_mask']
    df_std = analysis_result.get('df_std') # Use .get() for safety

    # Create a figure with a 3x2 grid
    fig, axs = plt.subplots(3, 2, figsize=(20, 24))
    fig.suptitle('Spring Detailed Analysis', fontsize=16)

    # 1. Local Coords vs Index
    ax1 = axs[0, 0]
    ax1.plot(local_coords[:, 0], label='Local X (Primary Axis)', alpha=0.8)
    ax1.plot(local_coords[:, 1], label='Local Y', alpha=0.8)
    ax1.plot(local_coords[:, 2], label='Local Z', alpha=0.8)
    ax1.set_title('1. Local Coords vs Index (Sorted & Oriented)')
    ax1.set_xlabel('Point Index')
    ax1.set_ylabel('Coordinate Value')
    ax1.legend()
    ax1.grid(True)

    # 2. Radius / Z vs Unwrapped Angle
    ax2 = axs[0, 1]
    ax2.plot(theta_unwrapped_deg, z, label='Z (Primary Axis)', color='tab:orange')
    ax2.plot(theta_unwrapped_deg, r, label='Radius', color='tab:green')
    ax2_twin = ax2.twinx()
    ax2_twin.plot(theta_unwrapped_deg, theta_unwrapped_deg, label='Unwrapped Angle', color='tab:blue', linestyle='--')
    ax2.set_title('2. Radius / Z vs Unwrapped Angle')
    ax2.set_xlabel('Unwrapped Angle (degrees)')
    ax2.set_ylabel('Radius / Z value')
    ax2_twin.set_ylabel('Angle (degrees)')
    ax2.legend(loc='upper left')
    ax2_twin.legend(loc='upper right')
    ax2.grid(True)

    # 3. Resampled Y vs X
    ax3 = axs[1, 0]
    ax3.plot(resampled_points[:, 0], resampled_points[:, 1])
    ax3.set_title('3. Top-Down View (Resampled Y vs X)')
    ax3.set_xlabel('X Coordinate (Primary Axis)')
    ax3.set_ylabel('Y Coordinate')
    ax3.axis('equal')
    ax3.grid(True)

    # 4. Resampled Z vs X
    ax4 = axs[1, 1]
    ax4.plot(resampled_points[:, 0], resampled_points[:, 2])
    ax4.set_title('4. Side View (Resampled Z vs X)')
    ax4.set_xlabel('X Coordinate (Primary Axis)')
    ax4.set_ylabel('Z Coordinate')
    ax4.axis('equal')
    ax4.grid(True)

    # 5. 3D plot of the resampled, centered data
    ax5 = fig.add_subplot(3, 2, 5, projection='3d')
    ax5.plot(resampled_points[:, 0], resampled_points[:, 1], resampled_points[:, 2], label='Resampled Data')
    # Mark outliers if any
    if np.any(outlier_mask):
        outlier_points = resampled_points[outlier_mask]
        ax5.scatter(outlier_points[:, 0], outlier_points[:, 1], outlier_points[:, 2],
                    color='red', s=20, label='Outliers')
    ax5.set_title('5. 3D View (Resampled & Centered)')
    ax5.set_xlabel('X')
    ax5.set_ylabel('Y')
    ax5.set_zlabel('Z')
    ax5.axis('equal')
    ax5.legend()

    # 6. Plot from 'zero-1' sheet data vs 'std' sheet data
    ax6 = axs[2, 1]
    ax6.set_title("6. 'zero-1' vs 'std' Sheet Data (Top-Down)")
    if df_std is not None:
        ax6.plot(df_std['y'], df_std['z'], label="'zero-1' (Y vs Z)", alpha=0.8, marker='.', linestyle='-')
    if std_df is not None:
        ax6.plot(std_df['y'], std_df['z'], label="'std' (Y vs Z)", alpha=0.7, marker='x', linestyle='--')
    
    ax6.set_xlabel('Y Coordinate')
    ax6.set_ylabel('Z Coordinate')
    ax6.axis('equal')
    ax6.legend()
    ax6.grid(True)

    plt.tight_layout(rect=(0, 0.03, 1, 0.97))
    
    # Save the figure
    output_path = os.path.join(output_dir, 'spring_detailed_analysis.png')
    plt.savefig(output_path)
    plt.close(fig)
    
    logger = logging.getLogger()
    logger.info(f" - {output_path}")

def read_std_sheet(filepath):
    """Reads the 'std' sheet from a given Excel file."""
    try:
        df = pd.read_excel(filepath, sheet_name='std')
        return df
    except FileNotFoundError:
        logging.warning(f"STD file not found at {filepath}. Skipping comparison.")
        return None
    except Exception as e:
        logging.error(f"Error reading STD sheet: {e}")
        return None

def _select_start_end_by_strategy(points, axis_point, axis_direction, percent=5.0):
    """Selects start/end slices based on radius, returning masks for them."""
    projections, distances, _ = project_points_to_axis(points, axis_point, axis_direction)
    p_min, p_max = np.min(projections), np.max(projections)
    slice_width = (p_max - p_min) * (percent / 100.0)
    
    start_slice_mask = projections < p_min + slice_width
    end_slice_mask = projections > p_max - slice_width
    
    return start_slice_mask, end_slice_mask, projections

def sort_points_along_axis(points, axis_point, axis_direction):
    """Sorts points along the given axis."""
    projections, _, _ = project_points_to_axis(points, axis_point, axis_direction)
    sort_indices = np.argsort(projections)
    return points[sort_indices], sort_indices, projections[sort_indices]

def order_points_by_helical_parameter(points: np.ndarray, axis_point: np.ndarray, axis_direction: np.ndarray,
                                      logger: logging.Logger | None = None) -> tuple[np.ndarray, np.ndarray, np.ndarray]:
    """Order points along a helical parameter using unwrapped angle in the local YZ plane.
    Returns (ordered_points, order_indices, helical_param).
    This avoids small misorderings from projection-only sorting near coil ends.
    """
    if logger is None:
        logger = logging.getLogger()
    # Build a temporary local frame where local X is the axis
    local_tmp, _ = to_local_coordinates(points, axis_point, axis_direction, logger)
    # Compute angle in YZ plane
    ang = np.arctan2(local_tmp[:, 2], local_tmp[:, 1])  # atan2(z_local, y_local)
    ang_unwrap = np.unwrap(ang)
    # Ensure direction aligns with axial progression
    ex = axis_direction / np.linalg.norm(axis_direction)
    proj = (points - axis_point) @ ex
    if np.corrcoef(ang_unwrap, proj)[0, 1] < 0:
        ang_unwrap = -ang_unwrap
    order = np.argsort(ang_unwrap)
    return points[order], order, ang_unwrap[order]

def resample_curve(points, num_points=1000):
    """Resamples a 3D curve to have n uniformly spaced points."""
    if len(points) < 2:
        return np.array([])

    # Calculate cumulative distance along the curve
    distances = np.cumsum(np.sqrt(np.sum(np.diff(points, axis=0)**2, axis=1)))
    distances = np.insert(distances, 0, 0)
    
    total_length = distances[-1]
    if total_length == 0:
        return points

    new_distances = np.linspace(0, total_length, num_points)
    
    resampled = np.zeros((num_points, 3))
    for i in range(3):
        resampled[:, i] = np.interp(new_distances, distances, points[:, i])
        
    return resampled

def _determine_origin_and_startidx(local_coords):
    """Determines the start index. Simplified version."""
    if len(local_coords) == 0:
        return 0
    # For now, always start at the beginning of the sorted array
    start_idx = 0
    return start_idx

def _roll(arr, shift):
    """Rolls an array."""
    if len(arr) == 0:
        return arr
    return np.roll(arr, -shift, axis=0)

def compute_curvature(points):
    """
    Calculates the curvature of a 3D curve defined by points.
    Uses the formula: K = |P' x P''| / |P'|^3
    """
    if len(points) < 3:
        return np.zeros(len(points))

    # First derivative (velocity)
    p_prime = np.gradient(points, axis=0)
    
    # Second derivative (acceleration)
    p_double_prime = np.gradient(p_prime, axis=0)

    # Cross product of velocity and acceleration
    cross_product = np.cross(p_prime, p_double_prime, axis=1)
    
    # Magnitude of the cross product
    norm_cross_product = np.linalg.norm(cross_product, axis=1)
    
    # Magnitude of the first derivative
    norm_p_prime = np.linalg.norm(p_prime, axis=1)
    
    # Curvature
    # Add a small epsilon to avoid division by zero
    epsilon = 1e-12
    curvature = norm_cross_product / (norm_p_prime**3 + epsilon)
    
    return curvature

def plot_cylindrical_3d(points, r, theta, z, path):
    """Plots 3D cylindrical data."""
    fig = plt.figure(figsize=(8, 6))
    ax = fig.add_subplot(111, projection='3d')
    ax.plot(r * np.cos(theta), r * np.sin(theta), z)
    ax.set_title("Cylindrical 3D Plot")
    plt.savefig(path)
    plt.close(fig)

def theta_modes(local_std, start_index=0):
    """Calculates various theta modes from local coordinates."""
    if len(local_std) == 0:
        return {
            'r': np.array([]), 'z': np.array([]), 'theta_raw_rad': np.array([]),
            'theta_raw_unwrapped_rad': np.array([]), 'theta_start0_unwrapped_rad': np.array([])
        }
    r = np.hypot(local_std[:, 0], local_std[:, 1])
    z = local_std[:, 2]
    theta_raw_rad = np.arctan2(local_std[:, 1], local_std[:, 0])
    theta_raw_unwrapped_rad = np.unwrap(theta_raw_rad)
    
    start_angle = theta_raw_rad[start_index] if start_index < len(theta_raw_rad) else 0
    theta_start0_unwrapped_rad = np.unwrap(theta_raw_rad - start_angle)
    
    return {
        'r': r,
        'z': z,
        'theta_raw_rad': theta_raw_rad,
        'theta_raw_unwrapped_rad': theta_raw_unwrapped_rad,
        'theta_start0_unwrapped_rad': theta_start0_unwrapped_rad,
    }

def wrap_to_pi(x):
    """Wraps angles to [-pi, pi]."""
    return (x + np.pi) % (2 * np.pi) - np.pi

def enforce_positive(theta_rad):
    """Ensures theta is positive increasing."""
    if len(theta_rad) > 1 and theta_rad[-1] < theta_rad[0]:
        return -theta_rad
    return theta_rad

def optimize_seam_min_gap(points, axis_dir, origin):
    """Placeholder for seam optimization."""
    return points, 0

def trim_to_complete_turns(points, axis_dir, origin):
    """Placeholder for trimming to complete turns."""
    return points, (0, len(points))

def create_new_result_report_chart(excel_path, output_path):
    """Create a comprehensive result report chart based on Z:AT area data."""
    logger = logging.getLogger()
    
    # Read the Excel data
    try:
        df = pd.read_excel(excel_path, sheet_name='zero-1')
    except Exception as e:
        logger.warning(f"결과 보고서 차트를 생성하지 못했습니다 (시트 읽기 실패): {e}")
        return
    
    # Filter to only numeric columns and drop NaN rows for key columns
    numeric_cols = ['N_turn', 'O_radius_copy', 'Q_perp_diam', 'S_vec_diam',
                     'T_abs_z', 'U_rel_height', 'W_pitch', 'X_min_pitch', 'θ']
    # Ensure required columns exist; if not, skip chart generation gracefully
    missing = [c for c in numeric_cols if c not in df.columns]
    if missing:
        logger.warning(f"차트 생성 스킵: 누락된 컬럼 {missing}")
        return
    df_numeric = df[numeric_cols].dropna()
    
    # Extract key columns for the report (equivalent to Z:AT area)
    turns = df_numeric['N_turn'].values.astype(float)
    radius = df_numeric['O_radius_copy'].values.astype(float)
    perp_diam = df_numeric['Q_perp_diam'].values.astype(float)
    vec_diam = df_numeric['S_vec_diam'].values.astype(float)
    abs_z = df_numeric['T_abs_z'].values.astype(float)
    rel_height = df_numeric['U_rel_height'].values.astype(float)
    pitch = df_numeric['W_pitch'].values.astype(float)
    min_pitch = df_numeric['X_min_pitch'].values.astype(float)
    theta_wrapped = df_numeric['θ'].values.astype(float)
    
    # Calculate summary statistics
    total_height = float(np.max(rel_height) - np.min(rel_height))
    total_turns = float(np.max(turns) - np.min(turns))
    mean_radius = float(np.mean(radius))
    mean_perp_diam = float(np.mean(perp_diam[perp_diam > 0]) if np.any(perp_diam > 0) else 0)
    mean_vec_diam = float(np.mean(vec_diam[vec_diam > 0]) if np.any(vec_diam > 0) else 0)
    mean_pitch = float(np.mean(pitch[pitch > 0]) if np.any(pitch > 0) else 0)
    
    # Create figure with multiple subplots
    fig = plt.figure(figsize=(20, 16))
    fig.suptitle('Spring Analysis Result Report (Z:AT Data)', fontsize=16, fontweight='bold')
    
    # 1. Summary Metrics Table
    ax1 = fig.add_subplot(3, 4, 1)
    ax1.axis('off')
    summary_text = f"""
    SUMMARY METRICS

    Total Height: {total_height:.2f} mm
    Total Turns: {total_turns:.2f}
    Mean Radius: {mean_radius:.2f} mm
    Mean Perp. Diam: {mean_perp_diam:.2f} mm
    Mean Vec. Diam: {mean_vec_diam:.2f} mm
    Mean Pitch: {mean_pitch:.2f} mm
    Min Pitch: {np.min(min_pitch):.2f} mm
    """
    ax1.text(0.05, 0.95, summary_text, transform=ax1.transAxes, 
             fontsize=9, verticalalignment='top', fontfamily='monospace',
             bbox=dict(boxstyle='round', facecolor='lightblue', alpha=0.8))
    
    # 2. Radius vs Turns
    ax2 = fig.add_subplot(3, 4, 2)
    ax2.plot(turns, radius, 'b-', linewidth=2, label='Radius')
    ax2.set_xlabel('Turns')
    ax2.set_ylabel('Radius (mm)')
    ax2.set_title('Radius Profile')
    ax2.grid(True, alpha=0.3)
    ax2.legend()
    
    # 3. Diameter Comparison
    ax3 = fig.add_subplot(3, 4, 3)
    valid_perp = perp_diam > 0
    valid_vec = vec_diam > 0
    if np.any(valid_perp):
        ax3.plot(turns[valid_perp], perp_diam[valid_perp], 'r-', linewidth=2, label='Perp. Diam')
    if np.any(valid_vec):
        ax3.plot(turns[valid_vec], vec_diam[valid_vec], 'g-', linewidth=2, label='Vec. Diam')
    ax3.set_xlabel('Turns')
    ax3.set_ylabel('Diameter (mm)')
    ax3.set_title('Diameter Analysis')
    ax3.grid(True, alpha=0.3)
    ax3.legend()
    
    # 4. Height vs Turns
    ax4 = fig.add_subplot(3, 4, 4)
    ax4.plot(turns, rel_height, 'purple', linewidth=2, label='Relative Height')
    ax4.plot(turns, abs_z, 'orange', linewidth=2, label='Absolute Z')
    ax4.set_xlabel('Turns')
    ax4.set_ylabel('Height/Z (mm)')
    ax4.set_title('Height Profile')
    ax4.grid(True, alpha=0.3)
    ax4.legend()
    
    # 5. Pitch Analysis
    ax5 = fig.add_subplot(3, 4, 5)
    valid_pitch = pitch > 0
    if np.any(valid_pitch):
        ax5.plot(turns[valid_pitch], pitch[valid_pitch], 'brown', linewidth=2, marker='o', markersize=3, label='Pitch')
        ax5.axhline(y=mean_pitch, color='red', linestyle='--', linewidth=2, label=f'Mean: {mean_pitch:.2f} mm')
    ax5.set_xlabel('Turns')
    ax5.set_ylabel('Pitch (mm)')
    ax5.set_title('Pitch Distribution')
    ax5.grid(True, alpha=0.3)
    ax5.legend()
    
    # 6. Theta vs Turns
    ax6 = fig.add_subplot(3, 4, 6)
    ax6.plot(turns, theta_wrapped, 'cyan', linewidth=2, label='Theta (wrapped)')
    ax6.set_xlabel('Turns')
    ax6.set_ylabel('Theta (degrees)')
    ax6.set_title('Angular Position')
    ax6.grid(True, alpha=0.3)
    ax6.legend()
    
    # 7. Radius Distribution Histogram
    ax7 = fig.add_subplot(3, 4, 7)
    ax7.hist(radius, bins=30, alpha=0.7, color='skyblue', edgecolor='black')
    ax7.axvline(x=mean_radius, color='red', linestyle='--', linewidth=2, 
                label=f'Mean: {mean_radius:.2f}')
    ax7.set_xlabel('Radius (mm)')
    ax7.set_ylabel('Frequency')
    ax7.set_title('Radius Distribution')
    ax7.legend()
    
    # 8. Pitch Distribution Histogram
    ax8 = fig.add_subplot(3, 4, 8)
    if np.any(valid_pitch):
        ax8.hist(pitch[valid_pitch], bins=20, alpha=0.7, color='lightgreen', edgecolor='black')
        ax8.axvline(x=mean_pitch, color='red', linestyle='--', linewidth=2,
                    label=f'Mean: {mean_pitch:.2f}')
    ax8.set_xlabel('Pitch (mm)')
    ax8.set_ylabel('Frequency')
    ax8.set_title('Pitch Distribution')
    ax8.legend()
    
    # 9. 3D Scatter Plot of Key Metrics
    ax9 = fig.add_subplot(3, 4, 9, projection='3d')
    scatter = ax9.scatter(radius.tolist(), turns.tolist(), rel_height.tolist(), c=theta_wrapped.tolist(), cmap='viridis', alpha=0.6)
    ax9.set_xlabel('Radius (mm)')
    ax9.set_ylabel('Turns')
    ax9.set_zlabel('Height (mm)')
    ax9.set_title('3D Parameter Space')
    plt.colorbar(scatter, ax=ax9, shrink=0.8, label='Theta (deg)')
    
    # 10. Correlation Matrix
    ax10 = fig.add_subplot(3, 4, 10)
    # Select numeric columns for correlation
    numeric_cols = ['N_turn', 'O_radius_copy', 'Q_perp_diam', 'S_vec_diam', 'T_abs_z', 'U_rel_height', 'W_pitch', 'θ']
    corr_data = df[numeric_cols].corr()
    im = ax10.imshow(corr_data, cmap='coolwarm', aspect='auto')
    ax10.set_xticks(range(len(numeric_cols)))
    ax10.set_yticks(range(len(numeric_cols)))
    ax10.set_xticklabels([col.split('_')[0] for col in numeric_cols], rotation=45)
    ax10.set_yticklabels([col.split('_')[0] for col in numeric_cols])
    ax10.set_title('Correlation Matrix')
    plt.colorbar(im, ax=ax10, shrink=0.8)
    
    # 11. Quality Metrics
    ax11 = fig.add_subplot(3, 4, 11)
    ax11.axis('off')
    quality_text = f"""
    QUALITY METRICS

    Data Points: {len(df)}
    Valid Pitch Points: {np.sum(valid_pitch)}
    Valid Diameters: {np.sum(valid_perp)}
    
    Radius Std Dev: {np.std(radius):.3f} mm
    Pitch Std Dev: {np.std(pitch[valid_pitch]):.3f} mm
    Height Range: {total_height:.2f} mm
    """
    ax11.text(0.05, 0.95, quality_text, transform=ax11.transAxes,
              fontsize=9, verticalalignment='top', fontfamily='monospace',
              bbox=dict(boxstyle='round', facecolor='lightgreen', alpha=0.8))
    
    # 12. Final Summary Plot - Combined metrics
    ax12 = fig.add_subplot(3, 4, 12)
    # Normalize and plot multiple metrics on same axis
    norm_turns = (turns - np.min(turns)) / (np.max(turns) - np.min(turns))
    norm_radius = (radius - np.min(radius)) / (np.max(radius) - np.min(radius))
    norm_height = (rel_height - np.min(rel_height)) / (np.max(rel_height) - np.min(rel_height))
    
    ax12.plot(norm_turns, norm_radius, 'b-', linewidth=2, label='Radius')
    ax12.plot(norm_turns, norm_height, 'r-', linewidth=2, label='Height')
    if np.any(valid_pitch):
        norm_pitch = (pitch[valid_pitch] - np.min(pitch[valid_pitch])) / \
                     (np.max(pitch[valid_pitch]) - np.min(pitch[valid_pitch]))
        ax12.plot(norm_turns[valid_pitch], norm_pitch, 'g-', linewidth=2, label='Pitch')
    
    ax12.set_xlabel('Normalized Turns')
    ax12.set_ylabel('Normalized Values')
    ax12.set_title('Normalized Profiles')
    ax12.legend()
    ax12.grid(True, alpha=0.3)
    
    plt.tight_layout()
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    plt.close()
    
    logger.info(f"New result report chart saved to: {output_path}")

def main():
    # Setup logging
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
    logger = logging.getLogger()

    # Determine the base directory (for PyInstaller)
    if getattr(sys, 'frozen', False):
        # Running as a bundled exe
        script_dir = os.path.dirname(sys.executable)
    else:
        # Running as a .py script
        script_dir = os.getcwd()

    # Determine IGS input file
    filepath = None
    # 1) If a command-line argument is provided, use it
    if len(sys.argv) > 1:
        arg = sys.argv[1]
        if os.path.isfile(arg) and arg.lower().endswith('.igs'):
            filepath = os.path.abspath(arg)
        else:
            logger.warning(f"지정된 인자 사용 불가 또는 확장자 불일치: {arg}")
    # 2) Search script directory
    if filepath is None:
        igs_files_here = [f for f in os.listdir(script_dir) if f.lower().endswith('.igs')]
        if igs_files_here:
            filepath = os.path.join(script_dir, igs_files_here[0])
    # 3) Search parent directory (useful when dist contains exe but data is one level up)
    if filepath is None:
        parent_dir = os.path.dirname(script_dir)
        if os.path.isdir(parent_dir):
            try:
                igs_files_parent = [f for f in os.listdir(parent_dir) if f.lower().endswith('.igs')]
                if igs_files_parent:
                    filepath = os.path.join(parent_dir, igs_files_parent[0])
            except Exception:
                pass
    # 4) As a last resort, search current working directory (if different)
    if filepath is None:
        cwd = os.getcwd()
        if cwd != script_dir:
            try:
                igs_files_cwd = [f for f in os.listdir(cwd) if f.lower().endswith('.igs')]
                if igs_files_cwd:
                    filepath = os.path.join(cwd, igs_files_cwd[0])
            except Exception:
                pass
    if filepath is None:
        logger.error(f".igs 파일을 찾을 수 없습니다. 실행 디렉토리: {script_dir}, 부모/작업폴더도 검색 실패")
        return None
    logger.info(f"입력 IGS 파일: {filepath}")

    # Excel workbook location: prefer same dir as IGS file
    igs_dir = os.path.dirname(filepath)
    output_dir = igs_dir
    output_excel_path = os.path.join(igs_dir, 'TK1.xlsx')
    
    # Check if TK1.xlsx exists; if not, try script_dir then parent
    if not os.path.exists(output_excel_path):
        alt_candidates = [
            os.path.join(script_dir, 'TK1.xlsx'),
            os.path.join(os.path.dirname(script_dir), 'TK1.xlsx'),
            os.path.join(os.getcwd(), 'TK1.xlsx')
        ]
        for alt in alt_candidates:
            if os.path.exists(alt):
                logger.info(f"TK1.xlsx 대체 위치 사용: {alt}")
                output_excel_path = alt
                output_dir = os.path.dirname(output_excel_path)
                break
        if not os.path.exists(output_excel_path):
            logger.error(f"TK1.xlsx 를 찾을 수 없습니다. IGS 폴더 또는 실행/부모 폴더에 배치하세요: {output_dir}")
            return None

    # --- 초기화: 프로그램 시작 시 zero-1 시트 A:X의 데이터(헤더 제외) 삭제 ---
    def clear_zero1_data(path: str, sheet: str):
        try:
            wb = load_workbook(path)
            if sheet not in wb.sheetnames:
                logger.warning(f"초기화 스킵: 시트 '{sheet}' 없음")
                return
            ws = wb[sheet]
            max_row = ws.max_row
            # 행 2 이후 A:X (1..24열) 값만 제거 (서식은 유지)
            for r in range(2, max_row + 1):
                for c in range(1, 25):
                    cell = ws.cell(row=r, column=c)
                    if cell.value is not None and str(cell.value).strip() != "":
                        cell.value = None
            safe_save_workbook(wb, path, logger)
            logger.info(f"zero-1 시트 데이터 초기화 완료 (A:X, 2행~{max_row}행)")
        except Exception as e:
            logger.warning(f"zero-1 시트 초기화 실패: {e}")

    clear_zero1_data(output_excel_path, 'zero-1')
    
    logger.info("IGS 파일 정밀 분석 시작...")
    original_points = parse_igs_points(filepath)
    logger.info(f"Loaded {len(original_points)} points")
    
    if len(original_points) == 0:
        logger.error("IGS 파일에서 포인트를 찾을 수 없습니다.")
        return None

    logger.info(f"총 데이터 포인트 수: {len(original_points)}")
    
    # --- PCA and Axis Definition ---
    center_of_mass, axis_vector, _, _ = find_spring_axis_pca(original_points)

    # =============================================================
    # Step 1, 2, 3 for Excel
    #   Step1: (x,y,z) = nearest original points mapped to standardized rows
    #   Step2: (x_norm,y_norm,z_norm) = 1,000-point arc-length resample (equidistant)
    #   Step3: (R,θ,Z) = cylindrical from Step2
    # Note: Full original points are exported separately as CSV.
    # =============================================================

    # Build an ordered path and resample to 1,000 points (equidistant by arc length)
    # Prefer helical-parameter ordering over pure axial projection for stability at ends
    points_sorted, _, _ = order_points_by_helical_parameter(original_points, center_of_mass, axis_vector, logger)
    resampled_points = resample_curve(points_sorted, num_points=1000)

    # Step2: local normalization from resampled points
    local_coords_step2, _ = to_local_coordinates(resampled_points, center_of_mass, axis_vector, logger)

    # Ensure primary axis direction is coherent for interpretability
    if len(local_coords_step2) > 0 and local_coords_step2[-1, 0] < local_coords_step2[0, 0]:
        local_coords_step2[:, 0] = -local_coords_step2[:, 0]
        local_coords_step2[:, 2] = -local_coords_step2[:, 2]
        logger.info("Standardization: Primary axis (X) direction flipped (resampled order).")

    # Normalize: start X from first; center Y/Z around mean
    if len(local_coords_step2) > 0:
        local_coords_step2[:, 0] -= local_coords_step2[0, 0]
        mean_yz_step2 = np.mean(local_coords_step2[:, 1:], axis=0)
        local_coords_step2[:, 1:] -= mean_yz_step2

    # Step3: cylindrical from Step2 (axis is local X; radial plane is YZ)
    cyl_Z = local_coords_step2[:, 0]
    cyl_Y = local_coords_step2[:, 1]
    cyl_X = local_coords_step2[:, 2]
    r_step3 = np.sqrt(cyl_Y**2 + cyl_X**2)
    theta_rad_step3 = np.arctan2(cyl_X, cyl_Y)
    theta_unwrapped_rad_step3 = np.unwrap(theta_rad_step3)
    theta_unwrapped_rad_step3 -= theta_unwrapped_rad_step3[0]
    theta_unwrapped_deg_step3 = np.rad2deg(theta_unwrapped_rad_step3)
    # Wrapped theta (display) in degrees, starting at 0 then confined to [-180,180]
    theta_wrapped_deg_step3 = np.rad2deg(wrap_to_pi(theta_unwrapped_rad_step3))

    # For quick inspection (optional pandas use)
    df_std = pd.DataFrame({
        'x': local_coords_step2[:, 0],
        'y': local_coords_step2[:, 1],
        'z': local_coords_step2[:, 2]
    })

    # --- Analysis pipeline (optional resampling for plots) ---
    local_coords_analysis, _ = to_local_coordinates(resampled_points, np.mean(resampled_points, axis=0), axis_vector, logger)
    if len(local_coords_analysis) > 0 and local_coords_analysis[-1, 0] < local_coords_analysis[0, 0]:
        local_coords_analysis[:, 0] = -local_coords_analysis[:, 0]
        local_coords_analysis[:, 2] = -local_coords_analysis[:, 2]
    start_idx = _determine_origin_and_startidx(local_coords_analysis) if len(local_coords_analysis) > 0 else 0
    local_coords_final = local_coords_analysis.copy()
    if len(local_coords_final) > 0:
        local_coords_final[:, 0] -= local_coords_final[start_idx, 0]
        start_point_yz = local_coords_final[start_idx, 1:]
        angle_to_zero = -np.arctan2(start_point_yz[1], start_point_yz[0]) + np.pi / 2
        c, s = np.cos(angle_to_zero), np.sin(angle_to_zero)
        rot_mat_yz = np.array([[c, -s], [s, c]])
        local_coords_final[:, 1:] = (rot_mat_yz @ local_coords_final[:, 1:].T).T
        mean_yz_analysis = np.mean(local_coords_final[:, 1:], axis=0)
        local_coords_final[:, 1:] -= mean_yz_analysis

    # Cylindrical for analysis plot only
    cyl_z = local_coords_step2[:, 0]
    cyl_y = local_coords_step2[:, 1]
    cyl_x = local_coords_step2[:, 2]
    r = np.sqrt(cyl_y**2 + cyl_x**2)
    theta_rad = np.arctan2(cyl_x, cyl_y)
    theta_unwrapped_rad = np.unwrap(theta_rad)
    theta_unwrapped_rad -= theta_unwrapped_rad[0]
    theta_unwrapped_deg = np.rad2deg(theta_unwrapped_rad)

    # --- Curvature Calculation (on analysis coords if available, else step2) ---
    curvature = compute_curvature(local_coords_final if len(local_coords_final) > 0 else local_coords_step2)

    # --- Outlier Detection ---
    curvature_threshold = np.mean(curvature) + 3 * np.std(curvature)
    outlier_mask = curvature > curvature_threshold

    # --- Excel Formula Implementation (map each standardized row to nearest raw by arc length) ---
    if len(local_coords_step2) > 0:
        # Cumulative distances along raw path (points_sorted polyline)
        d_raw = np.sqrt(np.sum(np.diff(points_sorted, axis=0)**2, axis=1))
        cum_raw = np.concatenate(([0.0], np.cumsum(d_raw)))
        # Cumulative distances along standardized path
        d_std = np.sqrt(np.sum(np.diff(resampled_points, axis=0)**2, axis=1)) if len(resampled_points) > 1 else np.array([])
        cum_std = np.concatenate(([0.0], np.cumsum(d_std))) if len(d_std) else np.array([0.0])
        # For each standardized arc position, find nearest raw index
        nearest_indices = np.searchsorted(cum_raw, cum_std, side='left')
        nearest_indices = np.clip(nearest_indices, 0, len(points_sorted) - 1)
        # Improve tie-breaking by checking previous index when possible
        for i, j in enumerate(nearest_indices):
            if j > 0 and j < len(cum_raw):
                if abs(cum_raw[j] - cum_std[i]) >= abs(cum_std[i] - cum_raw[j-1]):
                    nearest_indices[i] = j - 1
        raw_mapped = points_sorted[nearest_indices]
    else:
        raw_mapped = original_points

    # L, M, N: Turn calculation
    theta_wrapped_deg = theta_wrapped_deg_step3
    col_L = np.zeros_like(theta_wrapped_deg)
    col_L[1:] = np.diff(theta_wrapped_deg)
    col_M = np.zeros_like(col_L)
    for i in range(1, len(col_L)):
        col_M[i] = col_M[i-1] + (col_L[i] if col_L[i] > 0 else col_L[i-1])
    col_N = theta_unwrapped_deg_step3 / 360.0  # Correct turn calculation on original sequence

    # P: Index for vectorial diameter (half turn back) - MATCH formula
    p_targets = col_N - 0.5
    pos = np.searchsorted(col_N, p_targets, side='left') - 1
    col_P = np.where(pos < 0, 0, pos + 2)  # 1-based row numbers

    # S: Vectorial Diameter (safe indexing)
    col_S = np.zeros(len(raw_mapped), dtype=float)
    valid_p = col_P > 0
    if np.any(valid_p):
        idx = (col_P[valid_p] - 2).astype(int)
        idx = np.clip(idx, 0, len(raw_mapped) - 1)
        diffs = raw_mapped[valid_p] - raw_mapped[idx]
        col_S[valid_p] = np.linalg.norm(diffs, axis=1)

    # T: IF(S=0, MAX(S), S) (vectorial diameter fallback per user formula)
    max_s = float(np.max(col_S)) if len(col_S) else 0.0
    col_T = np.where(col_S == 0, max_s, col_S)

    # U: Copy of Z (absolute axial position J)
    col_U = cyl_Z

    # V: Index for pitch (full turn back)
    v_targets = col_N - 1.0
    col_V = np.searchsorted(col_N, v_targets, side='right') - 1
    col_V[col_V < 0] = 0

    # O: Copy of Radius (Column H)
    col_O = r_step3.copy()

    # Q: Perpendicular Diameter - updated formula (safe indexing)
    col_Q = np.zeros_like(col_O, dtype=float)
    if np.any(valid_p):
        idx_q = (col_P[valid_p] - 2).astype(int)
        idx_q = np.clip(idx_q, 0, len(col_O) - 1)
        col_Q[valid_p] = col_O[valid_p] + col_O[idx_q]

    # R: Fallback for Vectorial Diameter
    max_q = np.max(col_Q) if len(col_Q) > 0 else 0
    col_R = np.where(np.isclose(col_Q, 0), max_q, col_Q)

    # W: Pitch
    col_W = np.zeros_like(col_U)
    valid_v = col_V > 0
    col_W[valid_v] = col_U[valid_v] - col_U[col_V[valid_v]]

    # X: Min Pitch
    min_pitch_val = np.min(col_W[col_W > 0]) if np.any(col_W > 0) else 0
    col_X = np.full_like(col_W, min_pitch_val)

    # W: Pitch

    # --- Output Generation ---
    ensure_dir(output_dir)
    


    # --- Final Excel Output ---
    # Step1: raw sheet (all original points)
    try:
        write_excel_raw(output_excel_path, 'raw', original_points)
        logger.info(" - Wrote raw sheet (Step1) with all original points")
    except Exception as e:
        logger.warning(f"raw 시트 기록 실패: {e}")

    # Step2/3: standardized sheet zero-1 (1000 points)
    extra_cols_ordered = {
        'L_delta_theta': col_L,
        'M_cum_theta': col_M,
        'N_turn': col_N,
        'O_radius_copy': col_O,
        'P_half_turn_idx': col_P,
        'Q_perp_diam': col_Q,
        'R_fallback_diam': col_R,
        'S_vec_diam': col_S,
        'T_abs_z': col_T,       # now vectorial diameter filled (IF S=0 THEN max(S) ELSE S)
        'U_rel_height': col_U,  # now absolute Z copy (J column)
        'V_full_turn_idx': col_V,
        'W_pitch': col_W,
        'X_min_pitch': col_X,
    }

    write_excel_tk1(
        path=output_excel_path,
        sheet_name='zero-1',
        std_points=local_coords_step2,
        raw_mapped=raw_mapped,
        r_std=r_step3,
        theta_std_unwrapped=theta_wrapped_deg_step3,
        z_std=cyl_Z,
        extra_cols=extra_cols_ordered,
    )
    logger.info(f" - Successfully wrote to {output_excel_path}")

    # --- Visualization ---
    logger.info("\n상세 분석 시각화 생성 중...")
    
    # Read the std sheet for comparison
    std_df = read_std_sheet(os.path.join(script_dir, 'TK1_std.xlsx'))

    analysis_result = {
        'points': original_points,
        'center_of_mass': center_of_mass,
        'axis_vector': axis_vector,
        'local_coords': local_coords_step2,
        'resampled_points': local_coords_final if len(local_coords_final) > 0 else local_coords_step2,
        'r': r_step3,
        'z': cyl_Z,
        'theta_unwrapped_deg': theta_unwrapped_deg_step3,
        'outlier_mask': outlier_mask,
        'df_std': df_std,
        'col_N': col_N, # Add Turn data for plotting
    }
    visualize_spring_analysis(analysis_result, output_dir=output_dir, std_df=std_df)
    
    # --- Generate New Result Report Chart ---
    logger.info("새로운 결과 보고서 차트 생성 중...")
    create_new_result_report_chart(output_excel_path, os.path.join(output_dir, 'new_result_report.png'))
    
    logger.info("분석 완료.")
    return analysis_result

if __name__ == "__main__":
    main()