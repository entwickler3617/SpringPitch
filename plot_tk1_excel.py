#!/usr/bin/env python3
"""
Read TK1.xlsx (sheet 'zero-1') and plot the requested columns:
- (x_std, y_std, z_std)
- (R, θ, Z)

Saves figures under ./output/:
- xstd_ystd_zstd_3d.png: 3D trajectory of standardized points
- xstd_ystd_zstd_vs_index.png: x_std, y_std, z_std vs index
- R_theta_Z_vs_index.png: R, θ, Z vs index
If ./TK1.xlsx is locked/absent, falls back to ./output/TK1.xlsx
"""
import os
from typing import List, Tuple
import numpy as np
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D  # noqa: F401
from openpyxl import load_workbook
import logging

ROOT = os.path.dirname(os.path.abspath(__file__))
OUT_DIR = os.path.join(ROOT, 'output')

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


def load_columns_xlsx(path: str, sheet: str, headers: List[str]) -> Tuple[np.ndarray, ...]:
    wb = load_workbook(path, data_only=True, read_only=True)
    if sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet}' not found in {path}")
    ws = wb[sheet]

    # Determine column indices by header row (row=1)
    name_to_col = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if isinstance(v, str):
            name_to_col[v.strip()] = c
    idxs = []
    for h in headers:
        if h not in name_to_col:
            raise ValueError(f"Header '{h}' not found in sheet '{sheet}'")
        idxs.append(name_to_col[h])

    # Collect rows until all are None
    rows = []
    for r in range(2, ws.max_row + 1):
        vals = []
        all_none = True
        for c in idxs:
            v = ws.cell(row=r, column=c).value
            if v is not None and v != "":
                all_none = False
            vals.append(v)
        if all_none:
            break
        rows.append(vals)

    arr = np.array(rows, dtype=float) if rows else np.zeros((0, len(headers)))
    return tuple(arr[:, i] for i in range(arr.shape[1]))


def ensure_out_dir():
    os.makedirs(OUT_DIR, exist_ok=True)


def plot_xstd_ystd_zstd(x_std: np.ndarray, y_std: np.ndarray, z_std: np.ndarray):
    # 3D trajectory
    fig = plt.figure(figsize=(8, 7))
    ax = fig.add_subplot(111, projection='3d')
    ax.plot(x_std, y_std, z_std, color='teal', lw=1.5)
    ax.set_title('Standardized trajectory (x_std, y_std, z_std)')
    ax.set_xlabel('x_std')
    ax.set_ylabel('y_std')
    ax.set_zlabel('z_std')
    ax.view_init(elev=22, azim=-55)
    plt.tight_layout()
    path = os.path.join(OUT_DIR, 'xstd_ystd_zstd_3d.png')
    plt.savefig(path, dpi=160, bbox_inches='tight')
    logger.info(f"Saved: {path}")
    plt.close(fig)

    # Vs index
    n = len(x_std)
    idx = np.arange(1, n+1)
    fig, axs = plt.subplots(3, 1, figsize=(10, 8), sharex=True)
    axs[0].plot(idx, x_std, color='tab:blue'); axs[0].set_ylabel('x_std')
    axs[1].plot(idx, y_std, color='tab:green'); axs[1].set_ylabel('y_std')
    axs[2].plot(idx, z_std, color='tab:orange'); axs[2].set_ylabel('z_std'); axs[2].set_xlabel('index')
    fig.suptitle('x_std, y_std, z_std vs index')
    fig.tight_layout(rect=(0, 0.03, 1, 0.95))
    path = os.path.join(OUT_DIR, 'xstd_ystd_zstd_vs_index.png')
    plt.savefig(path, dpi=160, bbox_inches='tight')
    logger.info(f"Saved: {path}")
    plt.close(fig)


def plot_r_theta_z(R: np.ndarray, theta: np.ndarray, Z: np.ndarray):
    n = len(R)
    idx = np.arange(1, n+1)
    fig, axs = plt.subplots(3, 1, figsize=(10, 8), sharex=True)
    axs[0].plot(idx, R, color='tab:purple'); axs[0].set_ylabel('R')
    # θ 컬럼은 이제 −180..180 (deg, wrapped)로 저장됨
    axs[1].plot(idx, theta, color='tab:red'); axs[1].set_ylabel('θ (deg, wrapped)')
    axs[2].plot(idx, Z, color='tab:gray'); axs[2].set_ylabel('Z'); axs[2].set_xlabel('index')
    fig.suptitle('R, θ, Z vs index')
    fig.tight_layout(rect=(0, 0.03, 1, 0.95))
    path = os.path.join(OUT_DIR, 'R_theta_Z_vs_index.png')
    plt.savefig(path, dpi=160, bbox_inches='tight')
    logger.info(f"Saved: {path}")
    plt.close(fig)


def main():
    ensure_out_dir()
    primary = os.path.join(ROOT, 'TK1.xlsx')
    fallback = os.path.join(OUT_DIR, 'TK1.xlsx')
    xlsx = primary if os.path.exists(primary) else fallback
    logger.info(f"Reading: {xlsx}")

    # Load standardized Cartesian
    x_std, y_std, z_std = load_columns_xlsx(xlsx, 'zero-1', ['x_std', 'y_std', 'z_std'])
    plot_xstd_ystd_zstd(x_std, y_std, z_std)

    # Load cylindrical
    R, theta, Z = load_columns_xlsx(xlsx, 'zero-1', ['R', 'θ', 'Z'])
    plot_r_theta_z(R, theta, Z)


if __name__ == '__main__':
    main()
